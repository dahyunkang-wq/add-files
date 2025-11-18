# -*- coding: utf-8 -*-
import io
import json
import re
import zipfile
import base64
from io import BytesIO
from pathlib import Path
# [FIX] 타입 힌트(Tuple, List 등) 및 openpyxl 스타일 모듈 임포트 추가
from typing import List, Dict, Any, Tuple
import unicodedata  # 한글 자모 조합(NFC)을 위해 추가

import pandas as pd
import streamlit as st
import streamlit.components.v1 as components

# openpyxl 및 스타일 관련 모듈 추가
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.colors import Color

# [FIX] ModuleNotFoundError 해결을 위해 RichText 임포트 제거
# from openpyxl.text.rich_text import RichText
# from openpyxl.cell.text import Text


# =============================================================================
#
# 스크립트 1 (Excel → JSON) 헬퍼 함수
#
# =============================================================================

def normalize_category_name(raw_key: str) -> str:
    key = raw_key.strip().lower()
    key = key.replace(":", "")
    key = re.sub(r"\s+", "_", key)

    mapping = {
        "language": "language",
        "languages": "language",
        "audio_processing": "audio_processing",
        "audio": "audio_processing",
        "data_handling": "data_handling",
        "data": "data_handling",
        "tools": "tools",
        "tool": "tools",
    }

    return mapping.get(key, "etc")


def split_items(text: str):
    if not isinstance(text, str):
        return []

    parts = re.split(r"[\n,\r,]+", text)
    cleaned = []
    for p in parts:
        p = p.strip()
        p = re.sub(r'^[\*\-\·\u2022]+\s*', "", p)
        if p:
            cleaned.append(p)
    return cleaned


def parse_tech_stack(raw_text: str):
    result = {
        "language": [],
        "audio_processing": [],
        "data_handling": [],
        "tools": [],
        "etc": [],
    }

    if not isinstance(raw_text, str) or not raw_text.strip():
        return result

    lines = re.split(r"[\r\n]+", raw_text)
    current_key = None
    etc_buffer = []

    for line in lines:
        if not line or not line.strip():
            continue

        line = re.sub(r'^[\*\-\·\u2022]+\s*', "", line).strip()
        if not line:
            continue

        if ":" in line:
            raw_key, value = line.split(":", 1)
            cat = normalize_category_name(raw_key)
            current_key = cat

            items = split_items(value)
            if cat in result:
                result[cat].extend(items)
            else:
                result["etc"].extend(items)
        else:
            items = split_items(line)
            if current_key and current_key in result:
                result[current_key].extend(items)
            elif current_key and current_key not in result:
                result["etc"].extend(items)
            else:
                etc_buffer.extend(items)

    if etc_buffer:
        result["etc"].extend(etc_buffer)

    for key in list(result.keys()):
        seen = set()
        unique_items = []
        for item in result[key]:
            item = item.strip()
            if not item:
                continue
            if item in seen:
                continue
            seen.add(item)
            unique_items.append(item)
        result[key] = unique_items

    return result


def clean_task_description(raw_text: str) -> str:
    if not isinstance(raw_text, str):
        raw_text = str(raw_text) if raw_text is not None else ""
    text = re.sub(r"\s+", " ", raw_text).strip()
    return text


def excel_to_json_records(df: pd.DataFrame):
    records = []

    start_row = 11  # 12행
    num_rows = df.shape[0]

    for i in range(start_row, num_rows):
        d_val = df.iloc[i, 3] if df.shape[1] > 3 else None
        e_val = df.iloc[i, 4] if df.shape[1] > 4 else None
        f_val = df.iloc[i, 5] if df.shape[1] > 5 else None

        def is_empty(v):
            if v is None:
                return True
            if isinstance(v, float) and pd.isna(v):
                return True
            if isinstance(v, str) and not v.strip():
                return True
            return False

        if is_empty(d_val) and is_empty(e_val) and is_empty(f_val):
            break

        task_name = "" if d_val is None else str(d_val).strip()
        task_description = clean_task_description(e_val)
        tech_stack = parse_tech_stack("" if f_val is None else str(f_val))

        records.append(
            {
                "task_name": task_name,
                "task_description": task_description,
                "tech_stack": tech_stack,
            }
        )

    return records


# =============================================================================
#
# 스크립트 2 (JSON → Excel) 헬퍼 함수
#
# =============================================================================

# ==========================
# 상수 / 경로
# ==========================
# Streamlit에서 __file__은 스크립트 경로를 올바르게 참조합니다.
APP_DIR = Path(__file__).parent
TEMPLATE_DIR = APP_DIR / "templates"
DEFAULT_TEMPLATE_NONTRACK = "Non Track_Paper Interview_상위조직명_직무명(포맷).xlsx"
DEFAULT_TEMPLATE_TRACK    = "Track_Paper Interview_상위조직명_직무명(포맷).xlsx"

# Non Track 쓰기 범위
TASK_START_ROW_NT, TASK_END_ROW_NT   = 5, 14    # Task: A(이름), C(설명)
SKILL_START_ROW_NT, SKILL_END_ROW_NT = 5, 11    # Skill: A/B/D/F

# Track 쓰기 범위 (규칙 동일)
TASK_ROW_START_T, TASK_ROW_END_T   = 5, 14
SKILL_ROW_START_T, SKILL_ROW_END_T = 5, 11
TASK_TEMPLATE_SHEET_T  = "Task"
SKILL_TEMPLATE_SHEET_T = "Skill"
TRACK_TITLE_RANGE_T    = "D1:D2"  # 트랙명 표기 영역

# ==========================
# 공통: 텍스트 정리(마커 제거)
# ==========================
# [cite: ...]
CITE_PATTERN = re.compile(r'\s*\[\s*cite\s*:\s*.*?\]\s*', flags=re.IGNORECASE | re.DOTALL)
# (Source ...)
SOURCE_PAREN_PATTERN = re.compile(r'\s*\(\s*source[^)]*\)\s*', flags=re.IGNORECASE)

def strip_markers(text: Any) -> str:
    """[cite: ...], (Source ...) 제거 + 공백 정리"""
    if text is None:
        return ""
    s = str(text)
    s = CITE_PATTERN.sub(" ", s)
    s = SOURCE_PAREN_PATTERN.sub(" ", s)
    s = re.sub(r"[ \t]+", " ", s).strip()
    return s

# ==========================
# 공통: 파일명 유틸
# ==========================
INVALID_WIN_CHARS = r'<>:"/\\|?*'
INVALID_WIN_PATTERN = re.compile(f"[{re.escape(INVALID_WIN_CHARS)}]+")

def sanitize_filename_component(s: str, fallback: str = "untitled") -> str:
    if not s:
        return fallback
    s = INVALID_WIN_PATTERN.sub(" ", s).strip().strip(".")
    return s if s else fallback

# ==========================
# Non Track 파서/로직
# ==========================
def title_tokens_nt(stem: str) -> List[str]:
    return [t.strip() for t in stem.split("_") if t.strip()]

def is_trailing_excluded_nt(token: str) -> bool:
    t = token.lower().replace(" ", "")
    return t in {"skill", "hc제외"}

def parse_org_role_from_filename_nt(filename: str) -> Tuple[str, str, str]:
    """{상위조직명} = 첫 토큰, {직무명} = 두 번째~끝(뒤에서 skill/HC 제외 제거), 표시/파일명 둘 다 '공백' 연결"""
    stem = Path(filename).stem
    toks = title_tokens_nt(stem)
    if not toks:
        return "unknown", "", ""
    org = toks[0]
    end = len(toks)
    while end > 1 and is_trailing_excluded_nt(toks[end - 1]):
        end -= 1
    role_tokens = toks[1:end] if end > 1 else toks[1:]
    role_display = " ".join(role_tokens)
    role_for_filename = " ".join(role_tokens)
    return org, role_display, role_for_filename

def with_wrap(cell):
    a = cell.alignment or Alignment()
    return Alignment(
        horizontal=a.horizontal,
        vertical=a.vertical,
        text_rotation=a.text_rotation,
        wrap_text=True,
        shrink_to_fit=a.shrink_to_fit,
        indent=a.indent
    )

def set_text(ws, coord: str, text: str, wrap: bool = True):
    cell = ws[coord]
    cell.value = text
    if wrap:
        cell.alignment = with_wrap(cell)

def load_json_from_txt_bytes(b: bytes) -> Dict[str, Any]:
    """TXT에 전후 텍스트가 섞여 있어도 {} 블록만 추출 시도"""
    txt = b.decode("utf-8-sig", errors="ignore")
    try:
        return json.loads(txt)
    except json.JSONDecodeError:
        start = txt.find("{")
        end = txt.rfind("}")
        if start != -1 and end != -1 and start < end:
            return json.loads(txt[start:end+1])
        raise

def collect_tasks_nt(obj: Dict[str, Any]) -> List[Dict[str, Any]]:
    return obj.get("tasks") or []

def iter_skills_nt(obj: Dict[str, Any]):
    skills = obj.get("skills") or []
    for item in skills:
        if isinstance(item, dict) and "skill" in item:
            s = item.get("skill") or {}
            name = s.get("name", "")
            definition = s.get("definition", "")
            stack = s.get("tech_stack", {})
            related = item.get("related_tasks") or s.get("related_tasks") or []
        else:
            s = item if isinstance(item, dict) else {}
            name = s.get("name", "")
            definition = s.get("definition", "")
            stack = s.get("tech_stack", {})
            related = s.get("related_tasks") or []
        yield {"name": name, "definition": definition, "tech_stack": stack, "related_tasks": related}

def normalize_list(val) -> List[str]:
    if val is None:
        return []
    if isinstance(val, (list, tuple, set)):
        return [str(x).strip() for x in val if str(x).strip()]
    s = str(val).strip()
    if not s:
        return []
    parts = []
    for chunk in s.replace(";", ",").replace("/", ",").split(","):
        chunk = chunk.strip()
        if chunk:
            parts.append(chunk)
    return parts

def extract_tech_lines_nt(tech_stack: Dict[str, Any]) -> str:
    if not isinstance(tech_stack, dict):
        tech_stack = {}
    lower_map = {str(k).lower(): v for k, v in tech_stack.items()}
    languages = normalize_list(lower_map.get("language") or lower_map.get("languages"))
    os_list   = normalize_list(lower_map.get("os") or lower_map.get("platform") or lower_map.get("operating_system"))
    tools     = normalize_list(lower_map.get("tools") or lower_map.get("tool"))
    lines = []
    if languages: lines.append(f"* language: {', '.join(languages)}")
    if os_list:   lines.append(f"* os: {', '.join(os_list)}")
    if tools:     lines.append(f"* tools: {', '.join(tools)}")
    return strip_markers("\n".join(lines))  # ← 마커 제거

def bullet_lines(items: List[str]) -> str:
    items = [str(i).strip() for i in items if str(i).strip()]
    return "\n".join(f"* {i}" for i in items)

def related_task_names_nt(related_tasks: List[Dict[str, Any]], task_id_to_name: Dict[str, str]) -> List[str]:
    names = []
    for rt in related_tasks or []:
        name = (rt.get("task_name") or "").strip()
        if not name:
            tid = (rt.get("task_id") or "").strip()
            if tid and tid in task_id_to_name:
                name = task_id_to_name[tid]
        if name:
            names.append(name)
    return names

def build_workbook_nontrack(template_bytes: bytes, org: str, role: str, data: Dict[str, Any]) -> BytesIO:
    """템플릿 서식 유지, 값만 주입"""
    wb = load_workbook(BytesIO(template_bytes))
    ws_task  = wb["Task"] if "Task" in wb.sheetnames else wb[wb.sheetnames[0]]
    ws_skill = wb["Skill"] if "Skill" in wb.sheetnames else wb[wb.sheetnames[1]]

    # Task
    set_text(ws_task, "B1", org) # B1, B2는 VBA 수정 함수에서 한글 교정됨
    set_text(ws_task, "B2", role)
    tasks = collect_tasks_nt(data)
    task_id_to_name = {}
    for t in tasks:
        tid = str(t.get("task_id") or "").strip()
        tname = str(t.get("task_name") or "").strip()
        if tid and tname:
            task_id_to_name[tid] = tname
    row = TASK_START_ROW_NT
    for t in tasks[: (TASK_END_ROW_NT - TASK_START_ROW_NT + 1) ]:
        set_text(ws_task, f"A{row}", str(t.get("task_name") or "").strip())
        set_text(ws_task, f"C{row}", str(t.get("task_description") or "").strip())
        row += 1
    for r in range(row, TASK_END_ROW_NT + 1):
        set_text(ws_task, f"A{r}", ""); set_text(ws_task, f"C{r}", "")

    # Skill
    set_text(ws_skill, "B1", org) # B1, B2는 VBA 수정 함수에서 한글 교정됨
    set_text(ws_skill, "B2", role)
    processed = 0
    max_rows = SKILL_END_ROW_NT - SKILL_START_ROW_NT + 1
    for s in iter_skills_nt(data):
        if processed >= max_rows: break
        r = SKILL_START_ROW_NT + processed
        rel_names = related_task_names_nt(s.get("related_tasks"), task_id_to_name)
        set_text(ws_skill, f"A{r}", bullet_lines(rel_names) if rel_names else "")
        set_text(ws_skill, f"B{r}", str(s.get("name") or "").strip())
        set_text(ws_skill, f"D{r}", strip_markers(s.get("definition")))
        set_text(ws_skill, f"F{r}", extract_tech_lines_nt(s.get("tech_stack")))
        processed += 1
    for r in range(SKILL_START_ROW_NT + processed, SKILL_END_ROW_NT + 1):
        for c in ("A","B","D","F"):
            set_text(ws_skill, f"{c}{r}", "")

    # --- VBA 스타일 적용 ---
    apply_vba_description_edits(wb)
    apply_vba_extra_borders_and_dims(wb)
    apply_vba_global_font(wb, "현대하모니 L")
    apply_vba_korean_fix_to_headers(wb) # B1, B2 한글 교정
    # --- ---

    bio = BytesIO(); wb.save(bio); bio.seek(0); return bio

def process_uploaded_txt_nontrack(uploaded_file, template_bytes: bytes):
    org, role_display, role_for_filename = parse_org_role_from_filename_nt(uploaded_file.name)
    safe_org  = sanitize_filename_component(org, "org")
    safe_role = sanitize_filename_component(role_for_filename, "role")
    out_name = f"Non Track_Paper Interview_{safe_org}_{safe_role}.xlsx"
    data = load_json_from_txt_bytes(uploaded_file.read())
    # build_workbook_nontrack 내부에서 VBA 스타일 적용
    wb_bytes = build_workbook_nontrack(template_bytes, org, role_display, data)
    return out_name, wb_bytes

# ==========================
# Track 파서/로직
# ==========================
def parse_org_and_job_from_filename_track(filename: str) -> Tuple[str, str]:
    """
    파일명에서 상위조직/직무:
    - {상위조직} = '_' split 첫 토큰
    - {직무} = 첫 토큰 제외 후, 끝에서 'skill'/'HC 제외' 제거, 나머지를 '_'로 결합(원문 규칙 유지)
    """
    stem = Path(filename).stem
    tokens = stem.split("_")
    if not tokens:
        return "", ""
    org = tokens[0].strip()

    def norm(t: str) -> str: return t.lower().replace(" ", "")
    tail = tokens[1:]
    while tail and norm(tail[-1]) in ("skill", "hc제외"):
        tail.pop()
    job = "_".join(tail).strip()
    return org, job

# ---- 트랙 유틸 ----
def ensure_wrap(ws, row: int, col: int, vertical: str = "center"):
    existing = ws.cell(row=row, column=col).alignment or Alignment()
    ws.cell(row=row, column=col).alignment = Alignment(
        horizontal=existing.horizontal,
        vertical=vertical,
        wrap_text=True,
        text_rotation=existing.text_rotation,
        shrink_to_fit=existing.shrink_to_fit,
        indent=existing.indent
    )

def ensure_merge(ws, cell_range: str):
    existing = {str(rng) for rng in ws.merged_cells.ranges}
    if cell_range not in existing:
        ws.merge_cells(cell_range)

def set_vertical_center_all(ws):
    max_r, max_c = ws.max_row, ws.max_column
    for row in ws.iter_rows(min_row=1, max_row=max_r, min_col=1, max_col=max_c):
        for cell in row:
            a = cell.alignment or Alignment()
            cell.alignment = Alignment(
                horizontal=a.horizontal,
                vertical="center",
                wrap_text=a.wrap_text,
                text_rotation=a.text_rotation,
                shrink_to_fit=a.shrink_to_fit,
                indent=a.indent
            )

def copy_sheet_by_template(wb, template_sheet_name: str, new_title: str):
    src = wb[template_sheet_name]
    new_ws = wb.copy_worksheet(src)
    new_ws.title = new_title
    # column widths
    for key, dim in src.column_dimensions.items():
        new_ws.column_dimensions[key].width = dim.width
    # row heights
    for idx, dim in src.row_dimensions.items():
        if dim.height:
            new_ws.row_dimensions[idx].height = dim.height
    # merges
    src_merges = [str(r) for r in src.merged_cells.ranges]
    new_merges = {str(r) for r in new_ws.merged_cells.ranges}
    for r in src_merges:
        if r not in new_merges:
            new_ws.merge_cells(r)
    return new_ws

# ---- 트랙 데이터 선택 ----
def select_tasks_for_track(all_tasks: List[Dict[str, Any]], track_name: str, limit: int) -> List[Dict[str, Any]]:
    sel = [t for t in (all_tasks or []) if ((t.get("track") or {}).get("name")) == track_name]
    return sel[:limit]

def get_skill_field(s: Dict[str, Any], key: str, default=None):
    """스킬 항목이 {'skill': {...}} 또는 평평한 dict 모두 지원"""
    if isinstance(s, dict) and "skill" in s and isinstance(s["skill"], dict):
        return s["skill"].get(key, default)
    return s.get(key, default)

def get_skill_related_tasks(s: Dict[str, Any]):
    if isinstance(s, dict) and "skill" in s:
        return s.get("related_tasks") or s["skill"].get("related_tasks") or []
    return s.get("related_tasks") or []

def get_skill_track(s: Dict[str, Any]) -> Dict[str, Any]:
    # 주로 최상위에 'track'이 온다고 가정
    return s.get("track") or {}

def select_skills_for_track(all_skills: List[Dict[str, Any]], track_name: str, track_code: str, limit: int) -> List[Dict[str, Any]]:
    matched = []
    for s in all_skills or []:
        tr = get_skill_track(s) or {}
        scope = s.get("track_scope")
        name_match = (tr.get("name") == track_name) or (tr.get("code") == track_code)
        if name_match:
            matched.append(s); continue
        if scope == "common":
            for rt in get_skill_related_tasks(s) or []:
                trt = (rt.get("track") or {})
                if (trt.get("name") == track_name) or (trt.get("code") == track_code):
                    matched.append(s); break
    # 중복 제거(스킬명 기준)
    uniq, seen = [], set()
    for s in matched:
        sk_name = (get_skill_field(s, "name") or "").strip()
        if sk_name and sk_name not in seen:
            seen.add(sk_name); uniq.append(s)
    # rank 오름차순, None은 뒤
    def rank_key(s):
        r = get_skill_field(s, "rank")
        return (r is None, r if r is not None else 10**9)
    uniq.sort(key=rank_key)
    return uniq[:limit]

# ---- 트랙 본문 가공 ----
def bullets_from_related_tasks(related_tasks: List[Dict[str, Any]], current_track_name: str) -> str:
    if not related_tasks: return ""
    names, seen = [], set()
    for rt in related_tasks:
        tname = (rt or {}).get("task_name")
        ttrack = ((rt or {}).get("track") or {}).get("name")
        if tname and (ttrack == current_track_name) and (tname not in seen):
            seen.add(tname); names.append(tname)
    return "\n".join(f"* {n}" for n in names)

def listify_tech_value(v) -> List[str]:
    if v is None: return []
    if isinstance(v, (list, tuple, set)):
        return [strip_markers(x) for x in v if str(x).strip()]
    # 문자열이면 구분자로 분리
    return [strip_markers(x.strip()) for x in re.split(r"[;,/]", str(v)) if x.strip()]

def bullets_from_tech_stack(tech_stack: Dict[str, Any]) -> str:
    tech_stack = tech_stack or {}
    lines = []
    for key in ("language", "os", "tools"):
        vals = tech_stack.get(key)
        items = listify_tech_value(vals)
        items = [x for x in items if x]  # 빈 문자열 제거
        if items:
            lines.append(f"* {key}: {', '.join(items)}")
    return "\n".join(lines)

# ---- 트랙 시트 쓰기 ----
def write_task_sheet(ws, org_name: str, job_name: str, track_name: str, tasks: List[Dict[str, Any]]):
    ws["B1"].value = org_name # B1, B2는 VBA 수정 함수에서 한글 교정됨
    ws["B2"].value = job_name
    ensure_merge(ws, TRACK_TITLE_RANGE_T)
    ws["D1"].value = track_name
    ws["D1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row = TASK_ROW_START_T
    for t in tasks:
        if row > TASK_ROW_END_T: break
        ws.cell(row=row, column=1).value = t.get("task_name") or ""
        desc = t.get("task_description") or ""
        ws.cell(row=row, column=3).value = desc
        ensure_wrap(ws, row, 3, vertical="center")
        row += 1
    set_vertical_center_all(ws)

def write_skill_sheet(ws, org_name: str, job_name: str, track_name: str, skills: List[Dict[str, Any]]):
    ws["B1"].value = org_name # B1, B2는 VBA 수정 함수에서 한글 교정됨
    ws["B2"].value = job_name
    ensure_merge(ws, TRACK_TITLE_RANGE_T)
    ws["D1"].value = track_name
    ws["D1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row = SKILL_ROW_START_T
    for s in skills:
        if row > SKILL_ROW_END_T: break
        # A: 유관업무(현재 트랙 기준)
        a_text = bullets_from_related_tasks(get_skill_related_tasks(s), current_track_name=track_name)
        ws.cell(row=row, column=1).value = a_text
        ensure_wrap(ws, row, 1, vertical="center")
        # B: 스킬명
        ws.cell(row=row, column=2).value = (get_skill_field(s, "name") or "")
        # D: 설명(마커 제거)
        d_text = strip_markers(get_skill_field(s, "definition"))
        ws.cell(row=row, column=4).value = d_text
        ensure_wrap(ws, row, 4, vertical="center")
        # F: tech_stack(language/os/tools) (마커 제거 포함)
        f_text = bullets_from_tech_stack(get_skill_field(s, "tech_stack") or {})
        ws.cell(row=row, column=6).value = f_text
        ensure_wrap(ws, row, 6, vertical="center")
        row += 1
    set_vertical_center_all(ws)

def build_workbook_track(template_bytes: bytes, org: str, job: str, data: Dict[str, Any]) -> BytesIO:
    wb = load_workbook(BytesIO(template_bytes))

    # 트랙 목록(meta.tracks 우선)
    tracks = []
    meta_tracks = (((data.get("meta") or {}).get("tracks")) or [])
    if meta_tracks:
        for idx, tr in enumerate(meta_tracks, start=1):
            tracks.append({"index": idx, "name": tr.get("track_name"), "code": tr.get("track_code")})
    else:
        seen, idx = set(), 1
        for t in data.get("tasks", []):
            tn = (t.get("track") or {}).get("name")
            tc = (t.get("track") or {}).get("code")
            if tn and (tn, tc) not in seen:
                tracks.append({"index": idx, "name": tn, "code": tc})
                seen.add((tn, tc)); idx += 1

    all_tasks  = data.get("tasks")  or []
    all_skills = data.get("skills") or []

    for tr in tracks:
        t_idx = tr["index"]; t_name = tr["name"]; t_code = tr.get("code")
        # Task 시트
        task_ws_title = f"트랙 {t_idx}_Task"
        task_ws = copy_sheet_by_template(wb, TASK_TEMPLATE_SHEET_T, task_ws_title)
        tasks_for_track = select_tasks_for_track(all_tasks, t_name, limit=(TASK_ROW_END_T - TASK_ROW_START_T + 1))
        write_task_sheet(task_ws, org_name=org, job_name=job, track_name=t_name, tasks=tasks_for_track)
        # Skill 시트
        skill_ws_title = f"트랙 {t_idx}_Skill"
        skill_ws = copy_sheet_by_template(wb, SKILL_TEMPLATE_SHEET_T, skill_ws_title)
        skills_for_track = select_skills_for_track(all_skills, t_name, t_code, limit=(SKILL_ROW_END_T - SKILL_ROW_END_T + 1))
        write_skill_sheet(skill_ws, org_name=org, job_name=job, track_name=t_name, skills=skills_for_track)

    # 원본 템플릿 Task/Skill 시트 제거(Description 등은 유지)
    for base in (TASK_TEMPLATE_SHEET_T, SKILL_TEMPLATE_SHEET_T):
        if base in wb.sheetnames:
            wb.remove(wb[base])

    # --- VBA 스타일 적용 ---
    apply_vba_description_edits(wb)
    apply_vba_extra_borders_and_dims(wb)
    apply_vba_global_font(wb, "현대하모니 L")
    apply_vba_korean_fix_to_headers(wb) # B1, B2 한글 교정
    # --- ---

    bio = BytesIO(); wb.save(bio); bio.seek(0); return bio

def process_uploaded_txt_track(uploaded_file, template_bytes: bytes):
    org, job = parse_org_and_job_from_filename_track(uploaded_file.name)
    safe_org = sanitize_filename_component(org, "org")
    safe_job = sanitize_filename_component(job, "job")
    out_name = f"Track_Paper Interview_{safe_org}_{safe_job}.xlsx"
    data = load_json_from_txt_bytes(uploaded_file.read())
    # build_workbook_track 내부에서 VBA 스타일 적용
    wb_bytes = build_workbook_track(template_bytes, org, job, data)
    return out_name, wb_bytes

# ==========================
# 순차(멀티) 다운로드
# ==========================
def render_sequential_downloads(results_bytes: Dict[str, bytes], height: int = 240):
    if not results_bytes:
        return
    items_html, hidden_links_html = [], []
    for fname, b in results_bytes.items():
        b64 = base64.b64encode(b).decode("utf-8")
        mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        data_uri = f"data:{mime};base64,{b64}"
        items_html.append(f"<li>{fname}</li>")
        hidden_links_html.append(
            f'<a class="dl-link" href="{data_uri}" download="{fname}" style="display:none;"></a>'
        )
    html = f"""
<div id="bulk-dl">
  <button id="btn-bulk" style="padding:0.6rem 1rem;font-size:1rem;">📥 전체 파일 순차 다운로드</button>
  <p style="margin:0.5rem 0 0.25rem 0;">브라우저에서 다중 다운로드 허용이 필요할 수 있습니다.</p>
  <ul style="margin-top:0.25rem;">{''.join(items_html)}</ul>
  {''.join(hidden_links_html)}
</div>
<script>
(function() {{
  const btn = document.getElementById('btn-bulk');
  btn.addEventListener('click', async () => {{
    const links = Array.from(document.querySelectorAll('#bulk-dl a.dl-link'));
    for (const a of links) {{
      a.click();
      await new Promise(r => setTimeout(r, 400));
    }}
  }});
}})();
</script>
"""
    components.html(html, height=height, scrolling=False)


# =============================================================================
#
# 스크립트 2: VBA 서식 적용 헬퍼 (신규 추가)
#
# =============================================================================

# --- VBA: APPLY_GLOBAL_FONT ---
def apply_vba_global_font(wb, font_name: str):
    """워크북의 모든 셀에 글로벌 폰트 적용"""
    if not font_name:
        return
    try:
        # global_font = Font(name=font_name) # 이 줄은 사용되지 않으므로 주석 처리
        for ws in wb.worksheets:
            # 모든 셀을 순회하는 것은 매우 느리므로,
            # openpyxl에서는 기본 폰트를 변경하는 것이 더 효율적일 수 있으나
            # VBA 로직(Cells.Font.Name)을 따르기 위해 기존 셀 스타일을 순회합니다.
            # 하지만 더 효율적인 방법으로 스타일 객체를 수정합니다.
            
            # 1. 워크북의 기본 폰트 변경 시도 (테마 폰트가 사용된 경우)
            if wb.loaded_theme:
                if wb.loaded_theme.font_scheme and wb.loaded_theme.font_scheme.majorFont and wb.loaded_theme.font_scheme.majorFont.latin:
                    wb.loaded_theme.font_scheme.majorFont.latin.typeface = font_name
                if wb.loaded_theme.font_scheme and wb.loaded_theme.font_scheme.minorFont and wb.loaded_theme.font_scheme.minorFont.latin:
                    wb.loaded_theme.font_scheme.minorFont.latin.typeface = font_name

            # 2. 이미 개별 스타일이 적용된 셀 폰트 변경
            for row in ws.iter_rows():
                for cell in row:
                    if cell.has_style:
                        # 기존 폰트 속성을 유지하며 이름만 변경
                        cell.font = Font(
                            name=font_name,
                            sz=cell.font.sz,
                            b=cell.font.b,
                            i=cell.font.i,
                            color=cell.font.color,
                            underline=cell.font.underline,
                            strike=cell.font.strike,
                            family=cell.font.family,
                            scheme=cell.font.scheme,
                            charset=cell.font.charset,
                            vertAlign=cell.font.vertAlign,
                            outline=cell.font.outline,
                            shadow=cell.font.shadow,
                            condense=cell.font.condense,
                            extend=cell.font.extend
                        )
                    # else:
                        # 스타일이 없는 셀에도 폰트 적용 (VBA의 Cells.Font.Name 동작과 유사)
                        # cell.font = Font(name=font_name) 
                        # -> 이 로직은 모든 셀의 폰트를 개별 지정하여 파일 크기가 커질 수 있음.
                        # -> 템플릿이 이미 기본 폰트를 '현대하모니 L'로 설정했다면 필요 없음.
                        # -> VBA의 동작을 가장 가깝게 흉내 낸 것은 위 'if cell.has_style:' 블록임.

    except Exception as e:
        print(f"Warning: Global font '{font_name}' 적용 실패: {e}")


# --- VBA: APPLY_KOREAN_FIX ---
def apply_vba_korean_fix_to_headers(wb):
    """
    모든 '...Task' 및 '...Skill' 시트의 B1, B2 셀 값에
    한글 자모 조합(NFC 정규화)을 적용합니다.
    """
    try:
        for ws in wb.worksheets:
            if ws.title.endswith("Task") or ws.title.endswith("Skill"):
                for cell_coord in ["B1", "B2"]:
                    cell = ws[cell_coord]
                    if cell.value and isinstance(cell.value, str):
                        # NFC 정규화를 통해 자모음을 조합합니다.
                        normalized_text = unicodedata.normalize('NFC', cell.value)
                        if normalized_text != cell.value:
                            cell.value = normalized_text
    except Exception as e:
        print(f"Warning: Korean header fix (NFC) 적용 실패: {e}")


# --- VBA: APPLY_DESCRIPTION_EDITS ---
def apply_vba_description_edits(wb):
    """Description 시트 B8, B15 텍스트/스타일/크기 적용"""
    try:
        if "Description" not in wb.sheetnames:
            return
        
        ws = wb["Description"]
        
        # B열 너비 120
        ws.column_dimensions["B"].width = 120
        
        # [FIX] RichText 대신 기본 Font 객체만 정의
        # 강조(빨간색, 굵게) 폰트
        # highlight_font = Font(color=Color(rgb="FF0000"), bold=True)
        # 기본 폰트 (스타일 초기화용)
        default_font = Font(color=Color(rgb="000000"), bold=False)

        # B8: Task 안내
        txtB8 = (
            "Task Sheet는 팀의 업무분장표를 기준으로, '수행하시는 일(Task)'을 1차로 정리한 내용입니다.\n"
            "실제 현업의 관점에서 정확하게 작성되었는지 검토 및 확인 부탁드립니다.\n\n"
            "[검토 방법]\n"
            "▶ 1단계: ""Task 명""(A열)의 내용을 확인해보시고, "
        )
        highlightB8_1 = "수정사항이 있을 경우 ""Task 명"" 수정안""(B열)에 수정안을 작성해주세요."
        txtB8_cont = (
            "\n  - "
        )
        highlightB8_2 = "수정사항이 없다면 공란으로 두세요."
        txtB8_cont2 = (
            "\n\n▶ 2단계: ""Task 설명""(C열)의 내용을 확인해보시고, "
        )
        highlightB8_3 = "수정사항이 있을 경우 ""Task 설명"" 수정안""(D열)에 수정안을 작성해주세요."
        txtB8_cont3 = (
            "\n  - 예시) OO 업무는 실제 보안 측면으로 포커싱하고 있는데, 본 내용은 안전관리 측면으로 기입되어 있어 수정 필요합니다. 실제 하는 일은 ""~~~"" 입니다."
            "\n  - "
        )
        highlightB8_4 = "수정사항이 없다면 공란으로 두세요."

        # B8 RichText 적용 -> [FIX] 일반 텍스트로 변경
        ws["B8"].value = (
            txtB8 + highlightB8_1 + txtB8_cont + highlightB8_2 +
            txtB8_cont2 + highlightB8_3 + txtB8_cont3 + highlightB8_4
        )
        # [FIX] RichText를 사용하지 않으므로, 셀 전체에 기본 폰트를 적용합니다.
        ws["B8"].font = default_font # 기본 폰트 적용
        ws["B8"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[8].height = 165 # 행 높이

        # B15: Skill 안내
        txtB15 = (
            "[검토 방법]\n\n"
            "▶ 1단계: ""스킬명""(B열)의 내용을 확인해보시고, "
        )
        highlightB15_1 = "수정사항이 있을 경우 ""스킬 명"" 수정안""(C열)에 수정안을 작성해주세요."
        txtB15_cont = (
            "\n  - "
        )
        highlightB15_2 = "수정사항이 없다면 공란으로 두세요."
        txtB15_cont2 = (
            "\n  - A열의 '유관업무'는 B/D열에 있는 스킬이 실제 업무에서 어떻게 쓰이는지 보여주는 예시입니다. 이를 참고하여 이 스킬이 내 직무와 얼마나 관련 있는지 검토해 주세요.\n\n"
            "▶ 2단계: ""스킬 설명""(D열)의 내용을 확인해보시고, "
        )
        highlightB15_3 = "수정사항이 있을 경우 ""스킬 설명"" 수정안""(E열)에 수정안을 작성해주세요."
        txtB15_cont3 = (
            "\n  - "
        )
        highlightB15_4 = "수정사항이 없다면 공란으로 두세요."
        txtB15_cont4 = (
            "\n\n▶ 3단계: 실제 사용중인 스택 검토하기\n"
            "1) ""테크 스택""(F열)에 나열된 테크 스택을 확인해보시고, "
        )
        highlightB15_5 = "수정사항이 있을 경우 ""테크 스택""(G열)에 사용하는 스택명을 작성해주세요."
        txtB15_cont5 = (
            "\n  - "
        )
        highlightB15_6 = "수정사항이 없다면 공란으로 두세요."

        # B15 RichText 적용 -> [FIX] 일반 텍스트로 변경
        ws["B15"].value = (
            txtB15 + highlightB15_1 + txtB15_cont + highlightB15_2 +
            txtB15_cont2 + highlightB15_3 + txtB15_cont3 + highlightB15_4 +
            txtB15_cont4 + highlightB15_5 + txtB15_cont5 + highlightB15_6
        )
        # [FIX] RichText를 사용하지 않으므로, 셀 전체에 기본 폰트를 적용합니다.
        ws["B15"].font = default_font # 기본 폰트 적용
        ws["B15"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[15].height = 165 # 행 높이

    except Exception as e:
        print(f"Warning: Description 시트 편집(VBA) 적용 실패: {e}")


# --- VBA: APPLY_EXTRA_BORDERS ---
def apply_vba_extra_borders_and_dims(wb):
    """...Task / ...Skill 시트에 추가 테두리 및 크기 적용"""
    try:
        # 모든 테두리 스타일 정의 (thin)
        thin_border_side = Side(style='thin', color='000000')
        all_borders = Border(
            left=thin_border_side,
            right=thin_border_side,
            top=thin_border_side,
            bottom=thin_border_side
        )

        for ws in wb.worksheets:
            if ws.title.endswith("Task"):
                # Task 시트: A16:B16 테두리
                for row in ws["A16:B16"]:
                    for cell in row:
                        cell.border = all_borders
                # 16행 높이 53
                ws.row_dimensions[16].height = 53
            
            elif ws.title.endswith("Skill"):
                # Skill 시트: D열 너비 60
                ws.column_dimensions["D"].width = 60
                
                # G4:G11 테두리
                for row in ws["G4:G11"]:
                    for cell in row:
                        cell.border = all_borders
                
                # A13 테두리
                ws["A13"].border = all_borders
                # B13 테두리
                ws["B13"].border = all_borders
                
                # 13행 높이 53
                ws.row_dimensions[13].height = 53

    except Exception as e:
        print(f"Warning: 추가 테두리(VBA) 적용 실패: {e}")


# =============================================================================
#
# Streamlit 메인 UI
#
# =============================================================================

st.set_page_config(page_title="Excel ↔ JSON 변환 도구", layout="wide")
st.title("🚀 Excel ↔ JSON 변환 도구")
st.write("두 가지 변환 도구를 탭으로 분리하여 제공합니다.")

tab1, tab2 = st.tabs([
    "🛠️ 도구 1: 엑셀 (D12:F) → JSON 변환기",
    "✨ 도구 2: TXT (JSON) → 엑셀 (양식 채우기)"
])


# --- 탭 1: 엑셀 (D12:F) → JSON 변환기 (스크립트 1) ---
with tab1:
    st.header("엑셀 (D12~F열) → JSON txt 변환기")
    st.write("특정 포맷의 엑셀 파일(12행, D/E/F열)을 읽어 JSON으로 변환합니다.")

    uploaded_files_s1 = st.file_uploader(
        "엑셀 파일(.xlsx, .xls)을 하나 이상 선택하세요",
        type=["xlsx", "xls"],
        accept_multiple_files=True,
        key="excel_uploader_s1"  # 탭 간 구분을 위한 고유 키
    )

    if uploaded_files_s1:
        all_json_strings = {}
        st.subheader("변환 결과 미리보기")

        for file in uploaded_files_s1:
            st.markdown(f"### 파일: **{file.name}**")

            try:
                # [FIX] pandas가 openpyxl을 사용하도록 engine 명시
                df = pd.read_excel(file, header=None, engine='openpyxl')
            except Exception as e:
                st.error(f"{file.name} 읽기 실패: {e}")
                continue

            records = excel_to_json_records(df)
            json_str = json.dumps(records, ensure_ascii=False, indent=2)

            all_json_strings[file.name] = json_str

            st.code(json_str, language="json")

            base_name = file.name.rsplit(".", 1)[0]
            st.download_button(
                label=f"📄 {file.name} → JSON txt 다운로드",
                data=json_str.encode("utf-8"),
                file_name=f"{base_name}.json.txt",
                mime="text/plain",
                key=f"dl_json_{file.name}" # 개별 버튼 고유 키
            )

        if len(all_json_strings) > 1:
            st.subheader("ZIP으로 한 번에 받기")

            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
                for fname, jstr in all_json_strings.items():
                    base_name = fname.rsplit(".", 1)[0]
                    zf.writestr(f"{base_name}.json.txt", jstr)

            zip_buffer.seek(0)
            st.download_button(
                label="🗜️ 모든 JSON txt 파일 ZIP 다운로드",
                data=zip_buffer,
                file_name="json_outputs.zip",
                mime="application/zip",
                key="dl_zip_s1" # 고유 키
            )
    else:
        st.info("이곳에서 엑셀 파일을 업로드하면 JSON으로 변환됩니다.")


# --- 탭 2: TXT (JSON) → 엑셀 (양식 채우기) (스크립트 2) ---
with tab2:
    st.header("TXT(JSON) → Excel 변환기")
    st.write("특정 포맷의 JSON이 담긴 TXT 파일을 업로드하면, Non-Track/Track 엑셀 템플릿을 채웁니다.")

    # 탭 2의 모드 선택
    mode_s2 = st.radio(
        "모드 선택", 
        options=["Non Track", "Track"], 
        horizontal=True, 
        key="mode_s2" # 고유 키
    )

    # 템플릿 설정 (사이드바 대신 Expander 사용)
    with st.expander("템플릿 설정 (필수)", expanded=True):
        tpl_upload_s2 = st.file_uploader(
            "템플릿 업로드 (.xlsx) — (선택)", 
            type=["xlsx"], 
            accept_multiple_files=False, 
            key="tpl_uploader_s2" # 고유 키
        )

        template_bytes_s2 = None # 템플릿 로딩 상태
        
        if mode_s2 == "Non Track":
            default_tpl_path_name = DEFAULT_TEMPLATE_NONTRACK
            tpl_label = DEFAULT_TEMPLATE_NONTRACK
        else:
            default_tpl_path_name = DEFAULT_TEMPLATE_TRACK
            tpl_label = DEFAULT_TEMPLATE_TRACK

        if tpl_upload_s2 is None:
            # 기본 템플릿 로드 시도
            try:
                # Streamlit 배포 환경에서는 상대 경로가 다를 수 있으므로,
                # 스크립트 위치 기준으로 경로를 잡습니다.
                script_dir = Path(__file__).parent
                default_tpl_path_abs = script_dir / TEMPLATE_DIR / default_tpl_path_name

                if default_tpl_path_abs.exists():
                    st.success(f"기본 템플릿 사용: {tpl_label}")
                    template_bytes_s2 = default_tpl_path_abs.read_bytes()
                else:
                    st.error(f"기본 템플릿을 찾을 수 없습니다: {default_tpl_path_abs}")
            except Exception as e:
                st.error(f"기본 템플릿 로드 오류: {e}")
        else:
            template_bytes_s2 = tpl_upload_s2.read()
            st.success(f"업로드한 템플릿 사용: {tpl_upload_s2.name}")

        st.divider()
        if mode_s2 == "Non Track":
            st.markdown(
                """
    **규칙 요약 — Non Track**
    - 파일명  
      - `{상위조직명}` = `_` 분할 첫 토큰  
      - `{직무명}` = 두 번째 토큰부터, 끝에서 `'skill'`, `'HC 제외'` 제거 → 공백 연결  
    - **VBA 스타일 적용**: 
      - `Description` 시트 `B8`, `B15` 텍스트/서식/크기 적용 (열B=120, 행8/15=165)
      - `Task` 시트: `A16:B16` 테두리, 행16 높이 53
      - `Skill` 시트: `D`열 너비 60, `G4:G11`/`A13`/`B13` 테두리, 행13 높이 53
      - 전역 폰트 '현대하모니 L' 적용, `Task`/`Skill` 시트 `B1`/`B2` 한글 자모 교정
                """
            )
        else:
            st.markdown(
                """
    **규칙 요약 — Track**
    - 파일명  
      - `{상위조직명}` = `_` 분할 첫 토큰  
      - `{직무명}` = 첫 토큰 제외 후, 끝에서 `'skill'`, `'HC 제외'` 제거 → **`_`로 결합**
    - **VBA 스타일 적용**: 
      - `Description` 시트 `B8`, `B15` 텍스트/서식/크기 적용 (열B=120, 행8/15=165)
      - `트랙 n_Task` 시트: `A16:B16` 테두리, 행16 높이 53
      - `트랙 n_Skill` 시트: `D`열 너비 60, `G4:G11`/`A13`/`B13` 테두리, 행13 높이 53
      - 전역 폰트 '현대하모니 L' 적용, `Task`/`Skill` 시트 `B1`/`B2` 한글 자모 교정
                """
            )
    
    st.divider()

    # 탭 2의 파일 업로더
    st.subheader("1) TXT(JSON) 파일 업로드")
    st.warning("⚠️ **주의:** 이 기능은 '도구 1'에서 생성된 JSON과 호환되지 않습니다. 'Non-Track/Track' 템플릿에 맞는 별도의 JSON(txt) 파일을 업로드해야 합니다.")
    
    uploaded_files_s2 = st.file_uploader(
        "여러 파일을 동시에 올릴 수 있습니다.", 
        type=["txt"], 
        accept_multiple_files=True, 
        key="txt_uploader_s2" # 고유 키
    )

    # 탭 2의 미리보기
    if uploaded_files_s2:
        st.write("**파일명 파싱 미리보기**")
        preview_s2 = []
        for f in uploaded_files_s2:
            if mode_s2 == "Non Track":
                org, role_display, role_for_filename = parse_org_role_from_filename_nt(f.name)
                out = f"Non Track_Paper Interview_{sanitize_filename_component(org)}_{sanitize_filename_component(role_for_filename)}.xlsx"
                preview_s2.append({"원본 파일": f.name, "상위조직명": org, "직무명": role_display, "생성될 엑셀": out})
            else:
                org, job = parse_org_and_job_from_filename_track(f.name)
                out = f"Track_Paper Interview_{sanitize_filename_component(org)}_{sanitize_filename_component(job)}.xlsx"
                preview_s2.append({"원본 파일": f.name, "상위조직명": org, "직무명(파일 규칙)": job, "생성될 엑셀": out})
        st.dataframe(preview_s2, use_container_width=True)

    # 탭 2의 실행 버튼
    run_s2 = st.button(
        "변환 실행", 
        type="primary", 
        disabled=not uploaded_files_s2, 
        key="run_s2" # 고유 키
    )

    # 탭 2의 세션 상태 (탭 1과 분리)
    if "results_data_s2" not in st.session_state:
        st.session_state["results_data_s2"] = {}
    if "errors_data_s2" not in st.session_state:
        st.session_state["errors_data_s2"] = []
    if "last_mode_s2" not in st.session_state:
        st.session_state["last_mode_s2"] = mode_s2

    # 탭 2의 실행 로직
    if run_s2 and uploaded_files_s2:
        if template_bytes_s2 is None: # 템플릿이 로드되었는지 확인
            st.error("템플릿을 찾을 수 없습니다. 템플릿을 업로드하거나 기본 템플릿 경로를 확인하세요.")
        else:
            results_s2: Dict[str, bytes] = {}
            errors_s2: List[str] = []
            with st.spinner("변환 중..."):
                for uf in uploaded_files_s2:
                    try:
                        if mode_s2 == "Non Track":
                            name, bio = process_uploaded_txt_nontrack(uf, template_bytes_s2)
                        else:
                            name, bio = process_uploaded_txt_track(uf, template_bytes_s2)
                        results_s2[name] = bio.getvalue()
                    except Exception as e:
                        errors_s2.append(f"{uf.name} → 실패: {e} (line: {e.__traceback__.tb_lineno if e.__traceback__ else 'N/A'})") # 오류 디버깅을 위해 라인 번호 추가
            st.session_state["results_data_s2"] = results_s2
            st.session_state["errors_data_s2"] = errors_s2
            st.session_state["last_mode_s2"] = mode_s2

    # 탭 2의 결과 렌더링
    results_data_s2: Dict[str, bytes] = st.session_state.get("results_data_s2", {})
    errors_data_s2: List[str] = st.session_state.get("errors_data_s2", [])
    last_mode_s2 = st.session_state.get("last_mode_s2", mode_s2)

    if results_data_s2:
        st.subheader("2) 변환 결과")
        col1, col2 = st.columns([2, 1])

        with col1:
            st.success(f"{len(results_data_s2)}개 파일 생성 완료 — 모드: {last_mode_s2}")
            for fname, b in results_data_s2.items():
                st.download_button(
                    label=f"⬇️ {fname} 다운로드",
                    data=b,
                    file_name=fname,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key=f"dl_excel_{fname}" # 고유 키
                )

        with col2:
            render_sequential_downloads(results_data_s2) # 순차 다운로드

    if errors_data_s2:
        st.warning("일부 파일 변환 중 오류가 발생했습니다.")
        for msg in errors_data_s2:
            st.write(f"• {msg}")
